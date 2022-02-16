import json
import copy
from openpyxl import Workbook
from mdutils.mdutils import MdUtils
from openpyxl.worksheet.table import Table, TableStyleInfo

def set_key_values(en, zh):
    en['department'] = zh['部门']
    en['duties'] = zh['岗位']
    en['working_time'] = zh['上班时间']
    en['lunch_break'] = zh['午休时长']
    en['overtime'] = zh['加班情况']
    en['housing_provident_fund'] = zh['公积金']
    en['year-end_bonus'] = zh['年终奖']
    en['trial_period'] = zh['试用期']
    en['workstation'] = zh['工位电脑情况']
    en['annual_leave'] = zh['年假']
    en['punch_in'] = zh['打卡情况']
    en['remarks'] = zh['其他备注']
    en['update_time'] = zh['更新时间']

def zh_to_en(data):
    _data = []
    for co in data:
        en_co = {}
        en_co['company'] = co['公司名称']
        en_co['address'] = co['地址']
        if type(co['部门']) is list:
            _data_sub = []
            for de in co['部门']:
                en_sub_co = {}
                set_key_values(en_sub_co, de)
                _data_sub.append(en_sub_co)
            en_co['department'] = _data_sub;
        else:
            set_key_values(en_co, co)
        _data.append(en_co)
        
    with open('info_en.json', 'w', encoding='utf-8') as f:
        json.dump(_data, f, ensure_ascii=False, indent=4)


def generate_markdown(data):
    _data = copy.deepcopy(data)
    mdFile = MdUtils(file_name='net_company_info', title='互联网公司情况简表')
    for co in _data:
        mdFile.new_paragraph("**" + co['公司名称'] + "**")
        mdFile.new_paragraph("地址：" + co['地址'])
        del co['公司名称']
        del co['地址']
        if type(co['部门']) is list:
            list_of_departments = co['部门']
            list_of_strings = []
            for key in list_of_departments[0].keys():
                list_of_strings.append(key)
            
            for de in list_of_departments:
                list_of_strings.extend(list(de.values()))
            
            mdFile.new_line()
            mdFile.new_table(columns=13, rows=len(list_of_departments)+1, text=list_of_strings, text_align='center')
            mdFile.new_line('___')
        else:
            list_of_strings = []
            for key in co.keys():
                list_of_strings.append(key)
            list_of_strings.extend(list(co.values()))
            mdFile.new_line()
            mdFile.new_table(columns=13, rows=2, text=list_of_strings, text_align='center')
            mdFile.new_line('___')
    mdFile.create_md_file()

def generate_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width=25
    ws.column_dimensions['B'].width=25
    ws.column_dimensions['C'].width=10
    ws.column_dimensions['D'].width=10
    ws.column_dimensions['E'].width=25
    ws.column_dimensions['F'].width=15
    ws.column_dimensions['G'].width=20
    ws.column_dimensions['H'].width=25
    ws.column_dimensions['I'].width=25
    ws.column_dimensions['J'].width=25
    ws.column_dimensions['K'].width=25
    ws.column_dimensions['L'].width=25
    ws.column_dimensions['M'].width=10

    index = 0
    for co in data:
        index += 1
        ws.merge_cells('A' + str(index) + ':B' + str(index))
        company = co['公司名称']
        ws['A' + str(index)] = company
        index += 1
        ws.merge_cells('A' + str(index) + ':B' + str(index))
        ws['A' + str(index)] = '地址：' + co['地址']
        del co['公司名称']
        del co['地址']
        index += 1
        end = index
        department = co['部门']
        if type(department) is list:
            end += len(department)
            ws.append(list(department[0].keys()))
            for de in department:
                ws.append(list(de.values()))
        else:
            end += 1
            ws.append(list(co.keys()))
            ws.append(list(co.values()))
        tab = Table(displayName=company, ref="A" + str(index) + ":M" + str(end))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        index = end

    wb.save("net_company_info.xlsx")

if __name__ == '__main__':
    f = open('info.json')
    data = json.load(f)
    zh_to_en(data)
    generate_markdown(data)
    generate_excel(data)
    f.close()


    
